"""Case + scorecard exports: Excel and PDF.

Filename convention: case_scorecard_ddmmyyyyhhmmss (IST).
"""
import io
from datetime import datetime, timezone, timedelta

from sqlalchemy.orm import Session

from app.db import models

IST = timezone(timedelta(hours=5, minutes=30))


def export_filename() -> str:
    return f"case_scorecard_{datetime.now(IST):%d%m%Y%H%M%S}"


def export_filename_for(prefix: str) -> str:
    return f"{prefix}_{datetime.now(IST):%d%m%Y%H%M%S}"


def build_table_export(title: str, headers: list, rows: list, fmt: str) -> tuple[bytes, str]:
    """Generic single-table export used by the activity logs."""
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = title[:30]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0F172A")
        for row in rows:
            ws.append([str(c) for c in row])
        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 60)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=14 * mm, bottomMargin=12 * mm,
                            title=title, author="VITA", creator="VITA")
    styles = getSampleStyleSheet()
    body = [[Paragraph(str(c), styles["BodyText"]) for c in row] for row in rows]
    t = Table([headers] + body, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    doc.build([Paragraph(title, styles["Title"]), Spacer(1, 4 * mm), t])
    return buf.getvalue(), "application/pdf"


def _collect(db: Session, case: models.Case) -> dict:
    """Everything both formats need, in one place."""
    process = db.query(models.ProcessTemplate).get(case.inferred_process_id) \
        if case.inferred_process_id else None
    sc = (db.query(models.Scorecard).filter(models.Scorecard.case_id == case.id)
          .order_by(models.Scorecard.version.desc()).first())
    codes = {t.id: t.code for t in db.query(models.DocTypeTemplate).all()}
    docs = db.query(models.Document).filter(models.Document.case_id == case.id).all()
    discrepancies = db.query(models.Discrepancy).filter(
        models.Discrepancy.case_id == case.id).all()
    runs = (db.query(models.CaseRun).filter(models.CaseRun.case_id == case.id)
            .order_by(models.CaseRun.run_no).all())

    field_rows = []
    for d in docs:
        latest: dict[str, models.ExtractedField] = {}
        for f in d.fields:
            cur = latest.get(f.field_name)
            if cur is None or (f.extraction_round or 1) > (cur.extraction_round or 1):
                latest[f.field_name] = f
        for f in latest.values():
            field_rows.append([codes.get(d.doc_type_id, "UNKNOWN"), f.field_name,
                               f.value_normalized or "", float(f.confidence or 0),
                               f.extraction_round or 1])

    return {
        "summary": [
            ["Case reference", case.ref_no or ""],
            ["Case name", case.name or ""],
            ["Status", case.status],
            ["Business process", process.name if process else "—"],
            ["Process confidence", f"{float(case.inference_confidence or 0):.0f}%"],
            ["Overall score", f"{float(sc.overall_score):.1f}%" if sc else "—"],
            ["Scorecard version", sc.version if sc else "—"],
            ["Auto-verified", sc.auto_verified_count if sc else 0],
            ["Needs review", sc.review_needed_count if sc else 0],
            ["Hard fails", sc.hard_fail_count if sc else 0],
            ["Total runs", len(runs)],
            ["Exported (IST)", datetime.now(IST).strftime("%d %b %Y %H:%M:%S")],
        ],
        "exec_summary": (sc.summary if sc else "") or "",
        "fields": field_rows,
        "discrepancies": [
            [x.severity, x.kind, x.title, x.resolution] for x in discrepancies
        ],
        "runs": [
            [r.run_no, r.trigger, r.note or "",
             r.started_at.strftime("%d %b %H:%M") if r.started_at else "",
             r.scorecard_version or "",
             _diff_text(r.field_diff)] for r in runs
        ],
    }


def _diff_text(diff: dict | None) -> str:
    if not diff:
        return ""
    parts = []
    for item in diff.get("added", []):
        parts.append(f"+ {item['field']} = {item['value']}")
    for item in diff.get("updated", []):
        parts.append(f"~ {item['field']}: {item['old']} -> {item['new']}")
    for item in diff.get("deleted", []):
        parts.append(f"- {item['field']} (was {item['old']})")
    return "; ".join(parts)


# ------------------------------------------------------------------ Excel

def _build_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0F172A")

    def sheet(name, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = head, fill
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 60)
        return ws

    sheet("Summary", ["Item", "Value"], data["summary"], first=True)
    if data["exec_summary"]:
        wb["Summary"].append([])
        wb["Summary"].append(["Executive summary", data["exec_summary"]])
    sheet("Extracted Fields", ["Document", "Field", "Value", "Confidence %", "Round"],
          data["fields"])
    sheet("Discrepancies", ["Severity", "Kind", "Title", "Resolution"],
          data["discrepancies"])
    sheet("Run History", ["Run", "Trigger", "Note", "Started", "Scorecard v", "Changes"],
          data["runs"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ PDF

def _build_pdf(data: dict, doc_title: str = "VITA Case Scorecard") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)

    buf = io.BytesIO()
    # document metadata (title/author) — shown by PDF viewers instead of "anonymous"
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            title=doc_title, author="VITA", creator="VITA")
    styles = getSampleStyleSheet()
    story = [Paragraph("VITA — Case Scorecard", styles["Title"]),
             Spacer(1, 4 * mm)]

    def table(title, headers, rows, widths=None):
        story.append(Paragraph(title, styles["Heading2"]))
        body = [[Paragraph(str(c), styles["BodyText"]) for c in row] for row in rows]
        t = Table([headers] + body, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)
        story.append(Spacer(1, 5 * mm))

    table("Summary", ["Item", "Value"], data["summary"], widths=[55 * mm, 115 * mm])
    if data["exec_summary"]:
        story.append(Paragraph("Executive summary", styles["Heading2"]))
        story.append(Paragraph(data["exec_summary"], styles["BodyText"]))
        story.append(Spacer(1, 5 * mm))
    if data["fields"]:
        table("Extracted fields", ["Document", "Field", "Value", "Conf %", "Round"],
              data["fields"], widths=[28 * mm, 38 * mm, 66 * mm, 18 * mm, 16 * mm])
    if data["discrepancies"]:
        table("Discrepancies", ["Severity", "Kind", "Title", "Resolution"],
              data["discrepancies"], widths=[20 * mm, 32 * mm, 84 * mm, 32 * mm])
    if data["runs"]:
        table("Run history (retry audit)",
              ["Run", "Trigger", "Note", "Started", "Scorecard v", "Changes"],
              data["runs"], widths=[12 * mm, 18 * mm, 34 * mm, 24 * mm, 20 * mm, 62 * mm])

    doc.build(story)
    return buf.getvalue()


def build_export(db: Session, case: models.Case, fmt: str) -> tuple[bytes, str, str]:
    """Returns (content, media_type, filename) — the same base filename is used
    for the download and as the PDF's internal document title."""
    data = _collect(db, case)
    fname = export_filename()
    if fmt == "xlsx":
        return (_build_xlsx(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                fname)
    return _build_pdf(data, doc_title=fname), "application/pdf", fname
